#!/usr/bin/env python
# coding=UTF-8
"""
@Author: STAURL.COM
@Contact: admin@staurl.com
@Project: python_full_stack_automation_test
@File: test_cw0625_excel_package_1.py
@Time: 2019-06-30 11:44
@Desc: S
"""

from openpyxl import load_workbook  # 导入openpyxl第三方库
from scripts.handle_config import do_config
from scripts.handle_path import DATA_TEST_COMMON_FILE_PATH


class HandleExcel:  # 创建一个Excel处理类
    def __init__(self, filename, sheetname=None):  # 初始化类，并设置两个参数
        self.filename = filename
        self.sheetname = sheetname
        self.wb = load_workbook(filename)

    def active_sheet(self):
        if self.sheetname is None:  # 判断是否有指定表单，如果没有使用打开Excel文件的激活表单，如果有就是用指定名称的表单
            ws = self.wb.active
        else:
            ws = self.wb[self.sheetname]
        return ws

    def get_cases(self):  # 获取行-测试用例
        ws = self.active_sheet()
        cases_list = []  # 创建一个空的用例列表
        cases_title_entity = tuple(ws.iter_rows(max_row=1, values_only=True))
        cases_title = cases_title_entity[0]
        for cases_entity in tuple(ws.iter_rows(min_row=2, values_only=True)):
            cases_list.append(dict(zip(cases_title, cases_entity)))
        return cases_list  # 将用例列表返回给当前函数

    def write_result(self, row, actual, result):
        ws = self.active_sheet()
        if isinstance(row, int) and (2 <= row <= ws.max_row):
            ws.cell(row=row, column=do_config.get_int('excel', 'actual_col'), value=actual)
            ws.cell(row=row, column=do_config.get_int('excel', 'result_col'), value=result)
            self.wb.save(self.filename)

    def close_excel(self):
        self.wb.close()


if __name__ == '__main__':  # 在当前路径下才可以运行
    file_name = DATA_TEST_COMMON_FILE_PATH  # 指定文件
    my_excel = HandleExcel(file_name, 'register')  # 实例化my_excel这个对象
    print(my_excel.get_cases())  # 将对象调用方法后的结果打印出来
    my_excel.write_result(2, 4, 5)
    my_excel.close_excel()
